
import os
import numpy as np
import pandas as pd
import pandapower as pp

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side

rename_sheet = {
#   'reXplan            : 'pandapower
    'generators'        : 'gen',
    'external_gen'      : 'ext_grid',
    'static_generators' : 'sgen',
    'transformers'      : 'trafo',
    'nodes'             : 'bus',
    'switches'          : 'switch',
    'loads'             : 'load',
    'cost'              : 'poly_cost',
    'lines'             : 'line',
}

rename_column = {
    'from_bus'      : 'from_bus',
    'to_bus'        : 'to_bus',
}

def style_formatting(ws):
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 3)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    font = Font(bold=True)
    border = Border(bottom=Side(border_style="thick"))
    for cell in ws[1]:
        cell.font = font
        cell.border = border

    font = Font(bold=True)
    border = Border(bottom=Side(border_style="thick"))
    for cell in ws[1]:
        cell.font = font
        cell.border = border

def rename_element(sheet, column, values, net, rename = False):
    # TODO: Implement code for rename option!
    if values.empty:
        pass

    elif values.dtype == bool:
        values = values.astype('object').map({True: 'True', False: 'False'})

    elif column == 'name':
        if sheet == 'ln_type':
            values = getattr(net, 'line')['std_type']
        elif sheet == 'tr_type':
            values = getattr(net, 'trafo')['std_type']
        elif rename or values.isna().any() or values.apply(lambda x: isinstance(x, (int, float))).all():
            if sheet == 'nodes':
                for number in values.index:
                    values[number] = 'bus' + str(number + 1)
            else:
                values.reset_index(drop=True, inplace=True)
                for number in values.index:
                    values[number] = rename_sheet[sheet] + str(number + 1)
    
    elif column == 'std_type':  # TODO: Matching to Names!
        if values.isna().any():
            values = getattr(net, rename_sheet[sheet])['std_type']
            values.reset_index(drop=True, inplace=True)
            for number in values.index:
                if sheet == 'lines':
                    values[number] = 'line_type' + str(number + 1)
                elif sheet == 'transformers':
                    values[number] = 'trafo_type' + str(number + 1)
            
    elif column == 'node' or column == 'node_p' or column == 'node_s' or column == 'from_bus' or column == 'to_bus': 
        if isinstance(net, pp.auxiliary.pandapowerNet):              
            bus_column = getattr(net, rename_sheet[sheet])[rename_column[column]]
            bus_names = net.bus.loc[bus_column.tolist(), 'name']
            values = bus_names.reset_index(drop=True)
            values = values.rename(column)
        else:
            raise TypeError('Provided datatype of network is not compliant')
    
    elif column == 'element':
        values = values.astype('object')
        element_type = getattr(net, rename_sheet[sheet])['et']
        for index in range(len(element_type)):
            if element_type.iloc[index] == 'l':
                values.at[index] = getattr(net, rename_sheet['lines'])['name'].loc[values.iloc[index]]
            elif element_type.iloc[index] == 'b':
                values.at[index] = getattr(net, rename_sheet['nodes'])['name'].loc[values.iloc[index]]
            elif element_type.iloc[index] == 't':
                values.at[index] = getattr(net, rename_sheet['transformers'])['name'].loc[values.iloc[index]]
            elif element_type.iloc[index] == 't3':
                values.at[index] = getattr(net, rename_sheet['transformers_3w'])['name'].loc[values.iloc[index]]
            else:
                raise ValueError("Given element type of switch is unknown.")
    else:
        pass    
        # print(f"No need to rename for: [{sheet}] - [{column}]") # For debugging
    return values

def subimport_profiles(profiles):
    if isinstance(profiles, dict):
        print("Heck yeah!")
    print("Heck no!")

def import_grid(net, rename = False, profiles = None): # Add profiles!

    """
	Creates a reXplan compliant network as excel file from pandapower.

    INPUT:
		net (dict) - pandapower format network
        rename (bool) - False: Naming as provided in pandapower net; True: Elements renamed with respective naming of network (not implemented yet)

    EXAMPLE:
		>>> import_grid(net)
		>>> import_grid(pn.case14(), rename=False)
    """
    # TODO: FOR profiles = VALUE:-----------------------------------
    # TODO: - Write data correctly in profiles

    # TODO: FOR rename = FALSE:-------------------------------------
    # TODO: - [profiles] add data, if missing, better Error Message!
    # TODO: - [lines] geodata missing
    # TODO: - [nodes] geodata handling for bus with multiple entries
    # TODO: - Better solution for necessary empty tabs in network.xlsx? -> use keys of rename_column? 

    # TODO: FOR rename = TRUE:--------------------------------------
    # TODO: - Validate same Functionality

    path = os.path.dirname(os.getcwd())
    fields_maps = pd.read_csv(os.path.join(path + "\\reXplan",'fields_map.csv'))
    dfs_dict = pd.read_excel('template.xlsx', sheet_name=None)

    for index, row in fields_maps.iterrows():
        value = row.iloc[2]
        key = row.iloc[0]
        if pd.notnull(value):
            if key != value:
                rename_column[key] = value

    for sheet in dfs_dict.keys():       
        if sheet == 'cost':
            df_cost = getattr(net, rename_sheet[sheet])
            df_cost = df_cost.sort_values(by='et')
            df_cost = df_cost.rename(columns={'et': 'type'})

            name_array = np.array([])
            for index in range(len(df_cost.index)):
                if df_cost.iloc[index].type == 'ext_grid':
                    from_sheet = 'external_gen'
                elif df_cost.iloc[index].type == 'gen':
                    from_sheet = 'generators'
                elif df_cost.iloc[index].type == 'sgen':
                    from_sheet = 'static_generators'
                name = dfs_dict[from_sheet].name[df_cost.iloc[index].element] # Extracted name!
                name_array = np.append(name_array, name)

            df_cost = df_cost.rename(columns={'element': 'name'})
            dfs_dict[sheet] = df_cost
            dfs_dict[sheet]['name'] = name_array         

        elif sheet == 'profiles':
            load_names = getattr(net, 'load')['name'].tolist()
            columns_profile = ['asset'] + load_names
            dfs_dict['profiles'] = pd.DataFrame(columns=columns_profile)
            dfs_dict['profiles'].loc[1] = ['field'] + ['max_p_mw'] * len(net.load)

        elif sheet == 'ln_type' or sheet == 'tr_type':
            if sheet == 'ln_type':
                var = 'line'
            else:
                var = 'trafo'
            columns = getattr(net, var).columns
            for column in columns:
                    if column in dfs_dict[sheet].keys():
                        old_values = getattr(net, var)[column]
                        values = rename_element(sheet, column, old_values, net, rename)
                        dfs_dict[sheet][column] = values.values # .values?
            dfs_dict[sheet] = dfs_dict[sheet].drop_duplicates()

        else:
            try:
                columns = getattr(net, rename_sheet[sheet]).columns
                for column in columns:
                    if column in dfs_dict[sheet].keys() and not (sheet == 'lines' and column == 'type'):
                        old_values = getattr(net, rename_sheet[sheet])[column]
                        values = rename_element(sheet, column, old_values, net, rename)
                        dfs_dict[sheet][column] = values.values # .values?

                    elif column in rename_column.values():
                        column_update = next((key for key, value in rename_column.items() if value == column), None)
                        if column_update is not None:
                            old_values = getattr(net, rename_sheet[sheet])[column]
                            values = rename_element(sheet, column_update, old_values, net, rename)
                            dfs_dict[sheet][column_update] = values.values

                    elif column == 'std_type':
                        old_values = getattr(net, rename_sheet[sheet])[column]
                        values = rename_element(sheet, column, old_values, net, rename)
                        dfs_dict[sheet]['type'] = values.values

                    else:
                        pass
                        #print(f"\nSheet: {rename_sheet[sheet]}; Column: {column} is NOT used in template sheet") # For Debugging
            except:
                # print(f"[{sheet},{column}]") # For Debugging
                pass

    if net.bus_geodata.index.max() == net.bus.index.max():
        net.bus_geodata = net.bus_geodata.sort_index()
        dfs_dict['nodes']['longitude'] = net.bus_geodata.x
        dfs_dict['nodes']['latitude'] = net.bus_geodata.y

    dfs_dict['network']['sn_mva'] = pd.Series(net.sn_mva)
    dfs_dict['network']['f_hz'] = pd.Series(net.f_hz)
    dfs_dict['network']['name'] = pd.Series(net.name)

    if profiles:
        subimport_profiles(profiles)

    wb = Workbook()
    wb.remove(wb['Sheet'])

    nec_sheet_names = ['switches', 'cost', 'tr_type', 'static_generators', 'transformers', 'generators']
    for sheet_name, df in dfs_dict.items():
        if not df.empty or sheet_name in nec_sheet_names: 
        # if not df.empty:
            ws = wb.create_sheet(sheet_name)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
        else:
            pass
        style_formatting(ws)

    wb.save('network.xlsx')
    print('\n- network.xlsx created successfully! - ')