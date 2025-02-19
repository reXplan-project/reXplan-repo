import warnings
import itertools

import pandas as pd
import numpy as np
import math as math

import matplotlib.pyplot as plt
import netCDF4 as nc
import pandapower as pp

from . import config
from . import engine

from . import fragilitycurve
from . import hazard

# For stratification using the R language
import rpy2.robjects as robjects
from rpy2.robjects.packages import importr
from rpy2.robjects import pandas2ri
import rpy2.rlike.container as rlc

from mpl_toolkits.basemap import Basemap

# For creating reXplan compliant network-files via pandapower
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side

# TODO: improve warning messages in Network and PowerElement
# TODO: displace fragility curve elements
# TODO: revise following contants
# OBS: Not all elements can be set to i_montecarlo = True. The montecarlo_database.csv file cannot be empty

TIMESERIES_CLASS = pd.Series

# Network element status
STATUS = {'on': 1, 'off': 0, 'reparing': -1, 'waiting': -2}

def build_class_dict(df, class_name):
	return {row.id: globals()[class_name](row.dropna(axis=0).to_dict()) for _, row in config.df_to_internal_fields(df).iterrows()}

def find_element_in_standard_dict_list(id, list):
	# return first element or None otherwise
	return next((x for x in list if 'id' in x and x['id'] == id), None)

def standard_dict(content, type, id, field):
	return {k: v for k, v in locals().items()}

def get_datatype_elements(object, class_):
	if isinstance(object, list):
		iterateOver = object
		out = []
	elif isinstance(object, dict):
		iterateOver = object.values()
		out = []
	elif hasattr(object, "__dict__"):
		iterateOver = vars(object).values()
		out = [standard_dict(content=object, type=type(object).__name__, id=object.id, field=key)
		for key, value in vars(object).items() if isinstance(value, class_)]
	else:
		iterateOver = []
		out = []
	return out + list(itertools.chain(*[get_datatype_elements(x, class_) for x in iterateOver]))

def build_database(standard_dict_list, get_value_from_content=True):
	columnNames = [*standard_dict_list[0]]
	columnNames.remove('content')
	formatted_dict = {}
	for d in standard_dict_list:
		value = d['content']
		if get_value_from_content:
			value = getattr(value, d['field'])
		formatted_dict[(*[d[x] for x in columnNames],)] = value
	out = pd.DataFrame.from_dict(formatted_dict)
	out.columns.names = columnNames
	return out
class Network:
	"""
	Class that is representing a power network for the simulation purposes.

	Attributes:
		id (int or None): Identifier for the network.
		f_hz (float or None): Frequency of the network in Hertz.
		sn_mva (float or None): Apparent power of the network in MVA.
		totalInstalledPower (float): Total installed power in the network.
		totalConventionalPower (float): Total conventional power in the network.
		totalRenewablePower (float): Total renewable power in the network.
		totalStoragePower (float): Total storage power in the network.
		nodes (dict): Dictionary of nodes in the network.
		generators (dict): Dictionary of generators in the network.
		staticGenerators (dict): Dictionary of static generators in the network.
		externalGenerators (dict): Dictionary of external generators in the network.
		loads (dict): Dictionary of loads in the network.
		transformers (dict): Dictionary of transformers in the network.
		transformerTypes (dict): Dictionary of transformer types in the network.
		lines (dict): Dictionary of lines in the network.
		lineTypes (dict): Dictionary of line types in the network.
		switches (dict): Dictionary of switches in the network.
		crews (dict): Dictionary of crews in the network.
		fragilityCurves (dict): Dictionary of fragility curves for the network.
		event (Hazard): Hazard event associated with the network.
		returnPeriods (dict): Dictionary of return periods for the network.
		pp_network (pandapowerNet or None): Pandapower network object.
		outagesSchedule (DataFrame or None): Schedule of outages in the network.
		crewSchedule (DataFrame or None): Schedule of crews in the network.
		switchesSchedule (DataFrame or None): Schedule of switches in the network.
		metrics (list): List of metrics for the network.
		mcVariables (list): List of Monte Carlo variables for the network.
		powerElements (dict): Dictionary of power elements in the network.
		calculationEngine (Engine): Calculation engine for the network.

	Methods:
		repackage_geodata(df): Repackages geodata columns in the DataFrame.
		create_pandapower_df(df): Creates a pandapower DataFrame from the given DataFrame.
		augment_element_with_typedata(df, df_type):	Augments elements with type data.
		node_name_to_id(df, cols=[], bus_ids={}, line_ids={}, tr_ids={}, gen_ids={}, sgen_ids={}, ext_grid_ids={}, load_ids={}, dcline_ids={}, storage_ids={}, element_col=[]): Converts node names to IDs.
		create_elements(df, create_function, network): Creates elements in the network.
		build_network_parameters(networkFile): Builds network parameters from the given file.
		build_nodes(networkFile): Builds nodes from the given file.
		build_generators(networkFile): Builds generators from the given file.
		build_loads(networkFile): Builds loads from the given file.
		build_transformers(networkFile): Builds transformers from the given file.
		build_lines(networkFile): Builds lines from the given file.
		build_switches(networkFile): Builds switches from the given file.
		build_crews(networkFile): Builds crews from the given file.
		allocate_profiles(networkFile):	Allocates profiles from the given file.
		build_network(networkFile, build_pp_network = True): Builds the network from the given file.
		update_failure_probability(intensity=None, ref_return_period=None):
		build_pp_network(df_network, df_bus, df_tr, df_tr_type, df_ln, df_ln_type, df_load, df_ex_gen, df_gen, df_sgen, df_switch, df_cost): Builds the pandapower network.
		get_failure_candidates(): Gets the failure candidates in the network.
		get_closest_available_crews(availableCrew, powerElements): Gets the closest available crews for the power elements.
		get_crews_traveling_time(crew, powerElements): Gets the traveling time for the crews.
		get_reparing_time(powerElementsID, powerElements): Gets the repairing time for the power elements.
		calculate_outages_schedule(simulationTime, hazardTime): Calculates the outages schedule for the network.
		get_switch_candidates(): Gets the switch candidates in the network.
		calculate_switches_schedule(simulationTime): Calculates the switches schedule for the network.
		get_powerelement(id): Gets the power element by ID.
		propagate_schedules_to_network_elements(): Propagates schedules to network elements.
		update_montecarlo_variables(standard_dict): Updates Monte Carlo variables.
		update_grid(montecarlo_database, debug=None): Updates the grid with Monte Carlo database.
		build_montecarlo_database(time): Builds the Monte Carlo database.
		build_timeseries_database(time): Builds the timeseries database.
		calculate_metrics(): Calculates the metrics for the network.
		build_metrics_database(): Builds the metrics database.
		run(time, debug=None, **kwargs): Runs the simulation for the given time period.
		calc_stratas(data, ref_rp, xmin, xmax, cv=0.1, maxStrata=10): Calculates the stratas for the given data.
	"""
	def __init__(self, simulationName):
		self.id = None
		self.f_hz = None
		self.sn_mva = None

		self.totalInstalledPower = 0	# TODO is used?
		self.totalConventionalPower = 0
		self.totalRenewablePower = 0
		self.totalStoragePower = 0

		self.nodes = {}
		self.generators = {}
		self.staticGenerators = {}
		self.externalGenerators = {}
		self.loads = {}
		self.transformers = {}
		self.transformerTypes = {}
		self.lines = {}
		self.lineTypes = {}
		self.switches = {}
		self.crews = {}

		self.fragilityCurves = fragilitycurve.build_fragility_curve_database(simulationName)
		self.event = hazard.Hazard(simulationName)
		self.returnPeriods = hazard.build_return_period_database(simulationName)

		self.pp_network = None

		self.outagesSchedule = None
		self.crewSchedule = None
		self.switchesSchedule = None

		self.metrics = []
		self.mcVariables = []

		self.build_network(config.path.networkFile(simulationName))
		self.calculationEngine = engine.pandapower(self.pp_network)

		self.powerElements = {
							**self.lines,
							**self.generators,
							**self.loads,
							**self.transformers,
							**self.nodes}
		print(f"Network for study case <{simulationName}> initialized.")

	def repackage_geodata(self, df):
		GEODATA = 'geodata'
		cols_to_merge = sorted([col for col in df.columns if GEODATA in col])
		merged_col = []
		for _, row in df.iterrows():
			if row[cols_to_merge].isnull().any(): 
				merged_col.append(None)
			else:
				if len(cols_to_merge) == 2:
					merged_col.append(tuple(row[cols_to_merge]))
				elif len(cols_to_merge) == 4:
					merged_col.append([list(row[cols_to_merge[0:2]]), list(row[cols_to_merge[2:]])])
				else:
					raise ValueError('Unexpected number of geodata columns')
		df[GEODATA] = merged_col
		df = df.drop(cols_to_merge, axis=1)
		return df
	
	def create_pandapower_df(self, df):
		df = df.replace({np.nan: None})
		df_pp = config.df_to_pandapower_object(df)
		df_pp = df_pp.loc[:, df_pp.columns.notna()]
		return df_pp
	
	def augment_element_with_typedata(self, df, df_type):
		STD_TYPE = 'std_type'
		df_type.rename(columns={config.COL_NAME_NAME: STD_TYPE}, inplace=True)
		df = df.merge(df_type, on=STD_TYPE)
		df = df.drop(STD_TYPE, axis=1)
		return df
	
	def node_name_to_id(
			self,
			df,
			cols=[],
			bus_ids={},
			line_ids={},
			tr_ids={},
			gen_ids={},
			sgen_ids={},
			ext_grid_ids={}, 
			load_ids={},
			dcline_ids={},
			storage_ids={},
			element_col=[]):
		for col in cols:
			df[col] = [bus_ids[row] for row in df[col]]
		if len(element_col) > 0:
			for index, row in df.iterrows():
				if row[element_col[1]] == 'b':
					row[element_col[0]] = bus_ids[row[element_col[0]]]
				elif row[element_col[1]] == 'l':
					row[element_col[0]] = line_ids[row[element_col[0]]]
				elif row[element_col[1]] in ['t', 't3']:
					row[element_col[0]] = tr_ids[row[element_col[0]]]
				elif row[element_col[1]] == 'gen':
					row[element_col[0]] = gen_ids[row[element_col[0]]]
				elif row[element_col[1]] == 'sgen':
					row[element_col[0]] = sgen_ids[row[element_col[0]]]
				elif row[element_col[1]] == 'ext_grid':
					row[element_col[0]] = ext_grid_ids[row[element_col[0]]]
				elif row[element_col[1]] == 'load':
					row[element_col[0]] = load_ids[row[element_col[0]]]
				elif row[element_col[1]] == 'dcline':
					row[element_col[0]] = dcline_ids[row[element_col[0]]]
				elif row[element_col[1]] == 'storage':
					row[element_col[0]] = storage_ids[row[element_col[0]]]
				else:
					raise ValueError('Unexpected switch element type')
				df.iloc[index] = row
		return df
	
	def create_elements(self, df, create_function, network):
		ids = {}
		for _, row in df.iterrows():
			kwargs = row.to_dict()
			kwargs['net'] = network
			if config.COL_NAME_NAME in df.columns:
				ids[row[config.COL_NAME_NAME]] = create_function(
								**{key: value for key, value in kwargs.items() if value is not None})
			else:
				create_function(**{key: value for key, value in kwargs.items() if value is not None})
		return ids
			
	def build_network_parameters(self, networkFile):
		df_network = pd.read_excel(networkFile, sheet_name=config.get_input_sheetname('Network'))
		for _, row in config.df_to_internal_fields(df_network).iterrows():
			for key, value in row.dropna(axis=0).to_dict().items():
				if hasattr(self, key):
					setattr(self, key, value)
				else:
					warnings.warn(f'Input parameter "{key}" unknown in network parameters.')
		return df_network

	def build_nodes(self, networkFile):
		df_nodes = pd.read_excel(networkFile, sheet_name=config.get_input_sheetname('Bus'))
		self.nodes = build_class_dict(df_nodes, 'Bus')
		return df_nodes

	def build_generators(self, networkFile):
		df_gen = pd.read_excel(networkFile, sheet_name=config.get_input_sheetname('Generator'))
		df_ex_gen = pd.read_excel(networkFile, sheet_name=config.get_input_sheetname('External Generator'))
		df_sgen = pd.read_excel(networkFile, sheet_name=config.get_input_sheetname('Static Generator'))
		self.generators = build_class_dict(df_gen, 'Generator')
		self.externalGenerators = build_class_dict(df_ex_gen, 'Generator')
		self.staticGenerators = build_class_dict(df_sgen, 'Generator')
		return df_gen, df_ex_gen, df_sgen

	def build_loads(self, networkFile):
		df_load = pd.read_excel(networkFile, sheet_name=config.get_input_sheetname('Load'))
		self.loads = build_class_dict(df_load, 'Load')
		return df_load

	def build_transformers(self, networkFile):
		df_transformers = pd.read_excel(
			networkFile, sheet_name=config.get_input_sheetname('Transformer'))
		df_tr_types = pd.read_excel(networkFile, sheet_name=config.get_input_sheetname('Transformer Type'))
		self.transformers = build_class_dict(df_transformers, 'Transformer')
		self.transformerTypes = build_class_dict(df_tr_types, 'Transformer')
		return df_transformers, df_tr_types

	def build_lines(self, networkFile):
		df_lines = pd.read_excel(networkFile, sheet_name=config.get_input_sheetname('Line'))
		df_ln_types = pd.read_excel(networkFile, sheet_name=config.get_input_sheetname('Line Type'))
		self.lines = build_class_dict(df_lines, 'Line')
		self.lineTypes = build_class_dict(df_ln_types, 'Line')
		return df_lines, df_ln_types

	def build_switches(self, networkFile):
		df_switches = pd.read_excel(
			networkFile, sheet_name=config.get_input_sheetname('Switch'))
		self.switches = build_class_dict(df_switches, 'Switch')
		return df_switches

	def build_crews(self, networkFile):
		df_crews = pd.read_excel(networkFile, sheet_name=config.get_input_sheetname('Crew'))
		self.crews = build_class_dict(df_crews, 'Crew')

	def allocate_profiles(self, networkFile):
		df_profiles = pd.read_excel(
			networkFile, sheet_name=config.get_input_sheetname('Profile'), header=[0, 1], index_col=0)
		df_profiles.reset_index(drop=True, inplace=True)
		for assetId, field in df_profiles:
			element = self.get_powerelement(assetId)
			if hasattr(element, field):  # element can be None depending on get_powerelement()
				setattr(element, field, df_profiles[(assetId, field)])
			else:
				warnings.warn(f'Missing "{field}" field or "{assetId}" asset during profile allocation.')

	def build_network(self, networkFile, build_pp_network=True):
		df_network = self.build_network_parameters(networkFile)
		df_nodes = self.build_nodes(networkFile)
		df_load = self.build_loads(networkFile)
		df_gen, df_ex_gen, df_sgen = self.build_generators(networkFile) 	# TODO: Try-Catch for Gen if missing -> sgen
		df_transformers, df_tr_types = self.build_transformers(networkFile)
		df_lines, df_ln_types = self.build_lines(networkFile)
		df_switches = self.build_switches(networkFile)
		df_cost = pd.read_excel(networkFile, sheet_name = config.get_input_sheetname('Cost'))
		self.build_crews(networkFile)
		self.allocate_profiles(networkFile)
		if build_pp_network:
			self.pp_network = self.build_pp_network(
				df_network=df_network,
				df_bus=df_nodes,
				df_tr=df_transformers,
				df_tr_type=df_tr_types,
				df_ln=df_lines,
				df_ln_type=df_ln_types,
				df_load=df_load,
				df_ex_gen=df_ex_gen,
				df_gen=df_gen,
				df_sgen=df_sgen, # TODO: DEBUG
				df_switch=df_switches,
				df_cost=df_cost
			)

	def update_failure_probability(self, intensity = None, ref_return_period = None):
		"""
		Updates the failure probability for all power elements.

		This method iterates through all power elements and updates their failure 
		probability based on the provided intensity and reference return period.

		Args:
			intensity (float, optional): The intensity value used to update the failure probability. If None, the default intensity is used.
			ref_return_period (float, optional): The reference return period for the failure probability calculation. If None, the default return period is used.

		Returns:
			dict: A dictionary of power elements with updated failure probabilities.
		"""	
		for el in self.powerElements.values():
			el.update_failure_probability(self, intensity=intensity, ref_return_period=ref_return_period)
		return self.powerElements

	def build_pp_network(
						self,
						df_network,
						df_bus,
						df_tr,
						df_tr_type,
						df_ln,
						df_ln_type,
						df_load,
						df_ex_gen,
						df_gen,
						df_sgen,
						df_switch,
						df_cost):
		# TODO: it seems this funciton is missplaced. Can it be moved to engine.pandapower?

		# Creating the empty network
		df_network_pp = self.create_pandapower_df(df_network)
		kwargs_network = df_network_pp.iloc[0].to_dict()
		pp_network = pp.create_empty_network(**{key: value for key, value in kwargs_network.items() if value is not None})

		# Creating the bus elements
		df_bus_pp = self.create_pandapower_df(df_bus)
		df_bus_pp = self.repackage_geodata(df_bus_pp)
		bus_ids = self.create_elements(df_bus_pp, pp.create_bus, pp_network)

		# Creating the transformer elements
		df_tr_pp = self.create_pandapower_df(df_tr)
		df_tr_type_pp = self.create_pandapower_df(df_tr_type)
		df_tr_pp = self.augment_element_with_typedata(df_tr_pp, df_tr_type_pp)
		df_tr_pp = self.node_name_to_id(df_tr_pp, ['lv_bus', 'hv_bus'], bus_ids)
		tr_ids = self.create_elements(df_tr_pp, pp.create_transformer_from_parameters, pp_network)

		# Creating the line elements
		df_ln_pp = self.create_pandapower_df(df_ln)
		df_ln_type_pp = self.create_pandapower_df(df_ln_type)
		df_ln_pp = self.augment_element_with_typedata(df_ln_pp, df_ln_type_pp)
		df_ln_pp = self.node_name_to_id(df_ln_pp, ['from_bus', 'to_bus'], bus_ids)
		df_ln_pp = self.repackage_geodata(df_ln_pp)
		line_ids = self.create_elements(df_ln_pp, pp.create_line_from_parameters, pp_network)

		# Creating the load elements
		df_load_pp = self.create_pandapower_df(df_load)
		df_load_pp = self.node_name_to_id(df_load_pp, ['bus'], bus_ids)
		load_ids = self.create_elements(df_load_pp, pp.create_load, pp_network)

		# Creating the external gen elements
		df_ex_gen_pp = self.create_pandapower_df(df_ex_gen)
		df_ex_gen_pp = self.node_name_to_id(df_ex_gen_pp, ['bus'], bus_ids)
		ex_gen_ids = self.create_elements(df_ex_gen_pp, pp.create_ext_grid, pp_network)

		# Creating the generator elements
		df_gen_pp = self.create_pandapower_df(df_gen)
		df_gen_pp = self.node_name_to_id(df_gen_pp, ['bus'], bus_ids)
		gen_ids = self.create_elements(df_gen_pp, pp.create_gen, pp_network)

		# Creating the static generator elements
		df_sgen_pp = self.create_pandapower_df(df_sgen)
		df_sgen_pp = self.node_name_to_id(df_sgen_pp, ['bus'], bus_ids)
		sgen_ids = self.create_elements(df_sgen_pp, pp.create_sgen, pp_network)

		# Creating the switch elements
		df_switch_pp = self.create_pandapower_df(df_switch)
		df_switch_pp = self.node_name_to_id(
										df_switch_pp,
										['bus'],
										bus_ids,
										line_ids = line_ids,
										tr_ids = tr_ids,
										element_col = ['element', 'et'])
		switches_ids = self.create_elements(df_switch_pp, pp.create_switch, pp_network) # For future implementation

		# Creating the cost function
		df_cost_pp = self.create_pandapower_df(df_cost)
		df_cost_pp = self.node_name_to_id(
										df_cost_pp,
										gen_ids=gen_ids,
										sgen_ids=sgen_ids,
										ext_grid_ids=ex_gen_ids,
										load_ids=load_ids,
										element_col=['element', 'et'])
		self.create_elements(df_cost_pp, pp.create_poly_cost, pp_network)
		return pp_network

	def get_failure_candidates(self):
		candidates = {}
		for x in [x for x in self.__dict__.values() if isinstance(x, dict)]:
			candidates.update({y.id:y for y in x.values() if hasattr(y, 'failureProb') and y.failureProb != None and y.i_montecarlo != True})
		priorities = [x.priority if x.priority  != None else float('inf') for x in candidates.values()] # infinity for None priorities
		return {x:candidates[x] for _,x  in sorted(zip(priorities, candidates))}

	def get_closest_available_crews(self, availableCrew, powerElements):
		aux = len(powerElements)
		return powerElements[0:aux], availableCrew[0:aux]

	def get_crews_traveling_time(self, crew, powerElements):
		# time = np.random.randint(10)
		time = 0
		return [time]*len(crew)

	def get_reparing_time(self, powerElementsID, powerElements):
		# TODO: correct formula for repairing time
		return [powerElements[x].normalTTR for x in powerElementsID]

	def calculate_outages_schedule(self, simulationTime, hazardTime):
		'''
		outagesSchedule = 1 if powerElement is available
		'''
		failureCandidates = self.get_failure_candidates()
		failureProbability = np.array([x.failureProb for x in failureCandidates.values()])
		randomNumber = np.random.rand(len(failureProbability))
		failure = np.where((randomNumber <= failureProbability), 
							np.random.randint(hazardTime.start, [hazardTime.stop]*len(failureCandidates)),
							simulationTime.stop)

		crewSchedule = pd.DataFrame([[1]*len(self.crews)] * simulationTime.maxduration,
									columns = self.crews.keys(),
									index = simulationTime.interval)

		outagesSchedule = pd.DataFrame([[STATUS['on']]*len(failureCandidates)] * simulationTime.maxduration,
									   columns = failureCandidates.keys(),
									   index = simulationTime.interval)

		for index, column in zip(failure, outagesSchedule):
			outagesSchedule[column].loc[index:] = STATUS['off']

		for index, row in outagesSchedule.iterrows():
			failureElements = row.index[row == 0].tolist()
			availableCrews = crewSchedule.loc[index][crewSchedule.loc[index] == 1].index.tolist()
			elementsToRepair, repairingCrews = self.get_closest_available_crews(availableCrews, failureElements)
			crewsTravelingTime = self.get_crews_traveling_time(repairingCrews, elementsToRepair)
			repairingTime = self.get_reparing_time(elementsToRepair, failureCandidates)
			# TODO Add user info for missing data that is necessary to execute code
			for t_0, t_1, e, c in zip(crewsTravelingTime, repairingTime, elementsToRepair, repairingCrews):
				if t_1 is None:
					raise ValueError("No time to repair (TTR) defined for power element that fails.")
				outagesSchedule.loc[index+1:index + t_0, e] = STATUS['waiting']
				outagesSchedule.loc[index+t_0+1:index + t_0 + t_1, e] = STATUS['reparing']
				# Following line can be removed if outagesSchedule is set to 1 on at failure time
				outagesSchedule.loc[index+t_0+t_1+1:, e] = STATUS['on']
				crewSchedule.loc[index+1:index+t_0+t_1, c] = e

			if not outagesSchedule.loc[index+1:].isin([STATUS['on']]).any().any():
				# print(outagesSchedule.join(crewSchedule))
				self.outagesSchedule = outagesSchedule
				self.crewSchedule = crewSchedule
				return

		self.outagesSchedule = outagesSchedule
		self.crewSchedule = crewSchedule

	def get_switch_candidates(self):
		return {k:v for k,v in self.switches.items() if v.associated_elements != None and v.i_montecarlo != True}

	def calculate_switches_schedule(self, simulationTime):
		"""
		Switches will negate 'closed_' state iif at least one of its disconnecting_element is in outagesSchedule and its value is zero. This is done by a XNOR between closed_ values and previous condition.
		switchesSchedule = 1 iif switch is closed
		OBS: If a powerElement is excluded from self.outagesSchedule (e.g. i_montecarlo = True) then it will play no role on the determination of its associated switches' status.
		"""
		switchesSchedule = pd.DataFrame.from_dict(
			{k: [v.closed_]*simulationTime.maxduration for k, v in self.get_switch_candidates().items()}, dtype='int')  # forcing int is important!
		
		switchesSchedule.index = simulationTime.interval
		for switch_id in switchesSchedule:
			filter = self.outagesSchedule.columns.intersection(self.switches[switch_id].associated_elements)
			switchesSchedule[switch_id] = (switchesSchedule[switch_id] == (self.outagesSchedule[filter] > 0).all(axis=1))*1 #XNOR
		self.switchesSchedule = switchesSchedule

	def get_powerelement(self, id):
		# It assumes ids are unique!
		return next((x[id] for x in [self.loads, self.generators, self.transformers, self.lines, self.switches, self.nodes] if id in x), None)

	def propagate_schedules_to_network_elements(self):
		"""
		Assigns 'in_service' field based on outagesSchedule for powerElements
		Assigns 'closed' field based on switchesSchedule for switches
		"""
		df_in_service = self.outagesSchedule > 0
		df_closed = self.switchesSchedule > 0
		iterator = [['in_service', df_in_service], ['closed', df_closed]]
		for field, y in iterator:
			for elementId in y:
				element = self.get_powerelement(elementId)
				setattr(element, field, y[elementId])
				self.update_montecarlo_variables(standard_dict(
					content = element,
					type = type(element).__name__,
					id = elementId,
					field = field))  # elements are pointers

	def update_montecarlo_variables(self, standard_dict):
		element = find_element_in_standard_dict_list(
			standard_dict['id'], self.mcVariables)
		if not element or element['field'] != standard_dict['field']:
			self.mcVariables.append(standard_dict)

	def update_grid(self, montecarlo_database, debug=None):
		for type, id, field in montecarlo_database.columns.dropna():  # useful for empty database
			if id == debug and field == 'in_service':
				print(f'before update_grid: {debug} is {self.get_powerelement(id).in_service}')
			setattr(self.get_powerelement(id), field,
					montecarlo_database[type, id, field])
			if id == debug and field == 'in_service':
				print(f'after update_grid {debug} is {self.get_powerelement(id).in_service}')

	def build_montecarlo_database(self, time):
		if not self.mcVariables:
			print("No montecarlo variables to build database.")
		return build_database(self.mcVariables).loc[time.start: time.stop-1]

	def build_timeseries_database(self, time):
		""" 
		From the list [loads, generators, transformers, lines, switches] it creates a database with all the fields corresponding to timeseries
		TODO: Replace list by a function that will recognise powerElement-type instances
		"""
		elements = [
			self.loads, self.generators, self.staticGenerators, self.externalGenerators,
			self.transformers, self.lines, self.switches, self.nodes
		]
		return build_database(get_datatype_elements(elements, TIMESERIES_CLASS)).loc[time.start: time.stop-1]  # TODO Check Time

	def calculate_metrics(self):
		# DEPRECATED
		self.metrics = []

		def elements_in_service():
			return self.outagesSchedule[self.outagesSchedule > 0].sum(axis=1)

		def crews_in_service():
			return (self.crewSchedule != 1).sum(axis=1)

		self.metrics.append(Metric('Network', 'crews_in_service', crews_in_service()))
		self.metrics.append(Metric('Network', 'elements_in_service', elements_in_service()))
		self.metrics.append(Metric('Network', 'total_elements_in_service', elements_in_service().sum(), subfield='subfield_1', unit='unit_1'))

	def build_metrics_database(self):
		# DEPRECATED
		# TODO: make it recurrent
		out = []
		for metric in self.metrics:
			keys, values = zip(*[(key, value) for key, value in metric.__dict__.items() if key != 'value'])
			if isinstance(metric.value, pd.DataFrame) or isinstance(metric.value, pd.Series):
				value = metric.value.values
				index = metric.value.index
			else:
				value = metric.value
				index = [0]
			df = pd.DataFrame(value, columns=pd.MultiIndex.from_tuples(
				[values], names=keys), index=index)
			out.append(df)
		return pd.concat(out, axis=1)

	def run(self, time, debug=None, **kwargs):
		# TODO: to include an argument to choose which elements will not be considered in the timeseries simulation.
		# For instance, running a simulation with/without switches
		return self.calculationEngine.run(self.build_timeseries_database(time), debug=debug, **kwargs)

	def calc_stratas(self, data, ref_rp, xmin, xmax, cv=0.1, maxStrata=10):
		if maxStrata == 1:
			return pd.DataFrame({'Allocation': [1], 'Lower_X1': [xmin], 'Upper_X1': [xmax]})
		else:
			stratify = importr('SamplingStrata')
			base = importr('base')
			pandas2ri.activate()

			ids = list(range(len(data)))
			domain = [1]*len(data)

			Y = ["id"]
			y = [ids]
			cv_ = ["DOM1"]
			CV_ = ["DOM"]

			fc_rp_list = []
			for _, value in self.powerElements.items():
				if value.fragilityCurve and value.return_period:
					temp = [value.fragilityCurve, value.return_period]
					if temp not in fc_rp_list:
						fc_rp_list.append(temp)
			if not fc_rp_list:
				raise ValueError("No power element with fragility curve and return periods found. Check the power elements.")

			n = 1
			for fc_rp in fc_rp_list:
				Y.append(fc_rp[0]+'_'+fc_rp[1])
				y.append(self.fragilityCurves[fc_rp[0]].projected_fc(self.returnPeriods[fc_rp[1]], ref_rp, data))
				cv_.append(cv)
				CV_.append(str("CV" + str(n)))
				n += 1

			Y.extend(["x", "domain"])
			y.extend([data, domain])
			cv_.append(1)
			CV_.append("domainvalue")

			df = pd.DataFrame(np.transpose(np.array(y)), columns=Y)
			frame = stratify.buildFrameDF(
										df=robjects.conversion.py2rpy(df),
										id = "id",
										X = ["x"],
										Y = Y[1:-2],
										domainvalue = "domain")

			df_cv = base.as_data_frame(rlc.TaggedList(cv_, tags=tuple(CV_)))
			kmean = stratify.KmeansSolution2(
										frame = frame,
										errors = df_cv,
										maxclusters = maxStrata,
										showPlot = False) # TODO: Adjust UI Feedback
			try:
				if maxStrata == 1:
					nstrat = 1
				else:
					nstrat = np.unique(robjects.conversion.rpy2py(kmean)["suggestions"].values).size
				sugg = stratify.prepareSuggestion(kmean=kmean, frame=frame, nstrat=nstrat) # TODO: Adjust UI Feedback
				solution = stratify.optimStrata(
									method= "continuous",
									errors = df_cv,
                                    framesamp = frame,
									iter = 50,
                                    pops = 20,
									nStrata = nstrat,
									suggestions = sugg,
                                    showPlot = False,
                                    parallel = False) # TODO: Adjust UI Feedback (keep this!)
			except:
				warnings.warn(f'Cannot find optimal number of stratas. Try with a different reference return period. Continuing with 5 stratas.')
				solution = stratify.optimStrata(
									method= "continuous",
									errors = df_cv,
                                    framesamp = frame,
									iter = 50,
									pops = 20,
									nStrata = 5,
                                    showPlot = False,
                                    parallel = False) # CHECK THIS OUTPUT TOO
			strataStructure = stratify.summaryStrata(
										robjects.conversion.rpy2py(solution)[2],
                                        robjects.conversion.rpy2py(solution)[1],
                            			progress=False) # Why calling function 2 times?
			return (robjects.conversion.rpy2py(strataStructure))
class MonteCarloVariable:
	def __init__(self, element, id, field):
		self.element = element
		self.id = id
		self.field = field
class Metric:
	# TODO: delete
	def __init__(self, network_element, field, value, subfield=None, unit=None):
		self.network_element = network_element
		self.field = field
		self.value = value
		self.subfield = subfield
		self.unit = unit
class History:
	"""
	A class to represent the history of a network's performance metrics.

	Attributes:
		label (str): A label for the history instance.
		ENS (list): A list to store Energy Not Supplied (ENS) values.
		lineLoading (list): A list to store line loading values.
		minPower (int): The minimum power value.
		loadPower (int): The load power value.
		lineOutages (list): A list to store line outage events.
		transformerOutages (list): A list to store transformer outage events.
		busOutages (list): A list to store bus outage events.
		generatorOutages (list): A list to store generator outage events.
		LOEF (int): The Line Outage Event Frequency (LOEF).
		totENS (int): The total Energy Not Supplied (ENS).
		plot(): Placeholder method to plot the history data.
		export(): Placeholder method to export the history data.
		print(): Placeholder method to print the history data.

	Methods:
		__init__(label): Initializes the History object with the given label.
		plot():	Placeholder method to plot the history data.
		export(): Placeholder method to export the history data.
		print(): Placeholder method to print the history data.
	"""
	def __init__(self, label):
		self.label = label
		self.ENS = []
		self.lineLoading = []
		self.minPower = 0
		self.loadPower = 0
		self.lineOutages = []
		self.transformerOutages = []
		self.busOutages = []
		self.generatorOutages = []
		self.LOEF = 0
		self.totENS = 0

	def plot(self):
		pass

	def export(self):
		pass

	def print(self):
		pass
class GeoData:
	"""
	A class to represent geographical data with latitude and longitude.

	Attributes:
		latitude (float): The latitude of the geographical location.
		longitude (float): The longitude of the geographical location.

	Methods:
		__init__(latitude, longitude):
			Initializes the GeoData object with the given latitude and longitude.
	"""
	def __init__(self, latitude, longitude):
		self.latitude = latitude
		self.longitude = longitude
class PowerElement:
	"""
	A class to represent a power element in a network.

	Attributes:
		id (int): Unique identifier for the power element.
		node (int): Node identifier in the network.
		failureProb (float): Probability of failure of the power element.
		in_service (bool): Indicates if the power element is in service.
		fragilityCurve (str): Identifier for the fragility curve associated with the power element.
		return_period (float): Return period for the power element.
		normalTTR (float): Normal time to repair for the power element.
		priority (int): Priority of the power element.
		i_montecarlo (bool): Indicates if Monte Carlo simulation should be ignored.	# TODO CHECK IF BOOL OR INT

	Methods:
		__init__(**kwargs): Initializes the PowerElement with optional keyword arguments.
		update_failure_probability(network, intensity=None, ref_return_period=None): Updates the failure probability of the power element based on the network and intensity.
	"""
	def __init__(self, **kwargs):
		self.id = None
		self.node = None
		self.failureProb = None
		self.in_service = None
		self.fragilityCurve = None
		self.return_period = None
		self.normalTTR = None
		self.priority = None
		self.i_montecarlo = None

		for key, value in kwargs.items():
			if hasattr(self, key):
				setattr(self, key, value)
		if self.id is None:
			self.id = config.get_GLOBAL_ID()
			warnings.warn(f'No "id" defined as input for {self}. The following identifier was assigned: {self.id}.')
		if self.in_service == False:
			if self.i_montecarlo == False:
				warnings.warn(f'Forcing "ignore montecarlo" to "True" for {self.id}.')
			self.i_montecarlo = True

	def update_failure_probability(self, network, intensity=None, ref_return_period=None):
		if self.fragilityCurve == None:
			self.failureProb = None
		else:
			node = network.nodes[self.node]
			if intensity is None:
				if network.event.intensity is None:
					warnings.warn(f'Hazard event is not defined')
				else:
					_, event_intensity = network.event.get_intensity(node.longitude, node.latitude)
					intensity = event_intensity.max()
			if self.return_period != None and ref_return_period != None:
				self.failureProb = network.fragilityCurves[self.fragilityCurve].projected_fc(
					rp = network.returnPeriods[self.return_period],
					ref_rp = network.returnPeriods[ref_return_period],
					xnew = intensity)[0]
			else:
				self.failureProb = network.fragilityCurves[self.fragilityCurve].interpolate(intensity)[0]
class Bus(PowerElement):
	"""
	Represents a bus in a power network.

	Attributes:
		vn_kv (float): Voltage level in kV.
		min_vm_pu (float): Minimum voltage magnitude in per unit.
		max_vm_pu (float): Maximum voltage magnitude in per unit.
		zone (str): Zone in which the bus is located.
		type (str): Type of the bus.
		longitude (float): Longitude of the bus location.
		latitude (float): Latitude of the bus location.

	Methods:
		__init__(kwargs): Initializes the Bus with optional keyword arguments.
		update_failure_probability(network, intensity=None, ref_return_period=None): Updates the failure probability of the bus based on the given network, intensity, and reference return period.
	"""
	def __init__(self, kwargs):
		self.vn_kv = None
		self.min_vm_pu = None
		self.max_vm_pu = None
		self.zone = None
		self.type = None
		self.longitude = None
		self.latitude = None
		super().__init__(**kwargs)

	def update_failure_probability(self, network, intensity=None, ref_return_period=None):		
		if self.fragilityCurve == None:
			self.failureProb = None
		else:
			if intensity == None:
				_, event_intensity = network.event.get_intensity(self.longitude, self.latitude)
				intensity = event_intensity.max()
			if self.return_period != None and ref_return_period != None:
				self.failureProb = network.fragilityCurves[self.fragilityCurve].projected_fc(
					rp = network.returnPeriods[self.return_period],
					ref_rp = network.returnPeriods[ref_return_period],
					xnew = intensity)[0]
			else:
				self.failureProb = network.fragilityCurves[self.fragilityCurve].interpolate(intensity)[0]
class Switch(PowerElement):
	"""
	Initializes a Switch instance.

	A switch represents a connection between electrical elements like buses, lines, or transformers.
	The switch can be associated with specific elements and can have various attributes such as 
	current capacity, type, and state.

	Args:
		kwargs (dict): A dictionary of keyword arguments to initialize the switch attributes. 
			Expected keys include:
			- et (str): Element type ('l' = line, 't' = transformer, 't3' = transformer3w, 'b' = bus-to-bus).
			- element (str): The name or identifier of the connected element.
			- closed (bool): Indicates whether the switch is closed (True) or open (False).
			- in_ka (float): The rated current in kiloamperes.
			- type (str): The type of switch.
			- associated_elements (str): Comma-separated list of associated element names.

	Attributes:
		et (str): Type of element connection.
		element (str): Connected element name.
		closed (bool): Switch state.
		in_ka (float): Rated current in kA.
		type (str): Switch type.
		associated_elements (list): List of associated elements.
		closed_ (bool): Copy of the initial state of the switch.
	"""
	def __init__(self, kwargs):
		# element type: “l” = switch between bus and line, “t” = switch between bus and transformer, “t3” = switch between bus and transformer3w, “b” = switch between two buses'
		self.et = None
		self.element = None
		self.closed = None
		self.in_ka = None
		self.type = None
		self.associated_elements = None
		super().__init__(**kwargs)
		self.closed_ = self.closed
		if self.associated_elements:
			self.associated_elements = [
				x.strip() for x in self.associated_elements.split(',')]
class Generator(PowerElement):
	"""
	p_mw (float, optional): Active power in MW.
	q_mvar (float, optional): Reactive power in MVAr.
	vm_pu (float, optional): Voltage magnitude in per unit.
	controllable (bool, optional): Indicates if the generator is controllable.
	max_p_mw (float, optional): Maximum active power in MW.
	min_p_mw (float, optional): Minimum active power in MW.
	max_q_mvar (float, optional): Maximum reactive power in MVAr.
	min_q_mvar (float, optional): Minimum reactive power in MVAr.
	slack (bool, optional): Indicates if the generator is a slack bus.
	slack_weight (float, optional): Weight of the slack bus.
	sn_mva (float, optional): Apparent power in MVA.
	scaling (float, optional): Scaling factor.
	type (str, optional): Type of the generator.
	vn_kv (float, optional): Nominal voltage in kV.
	xdss_pu (float, optional): Subtransient reactance in per unit.
	rdss_ohm (float, optional): Subtransient resistance in ohms.
	cos_phi (float, optional): Power factor.
	pg_percent (float, optional): Percentage of active power generation.
	power_station_trafo (bool, optional): Indicates if the generator is connected to a power station transformer.
	va_degree (float, optional): Voltage angle in degrees.
	"""
	def __init__(self, kwargs):
		self.p_mw=None
		self.q_mvar=None
		self.vm_pu=None
		self.controllable=None
		self.max_p_mw=None
		self.min_p_mw=None
		self.max_q_mvar=None
		self.min_q_mvar=None
		self.slack=None
		self.slack_weight=None
		self.sn_mva=None
		self.scaling=None
		self.type=None
		self.vn_kv=None
		self.xdss_pu=None
		self.rdss_ohm=None
		self.cos_phi=None
		self.pg_percent=None
		self.power_station_trafo=None
		self.va_degree=None
		super().__init__(**kwargs)
class Load(PowerElement):
	"""
	Represents a load in a power network.

	Attributes:
		p_mw (float): Active power in megawatts.
		q_mvar (float): Reactive power in megavars.
		controllable (bool): Indicates if the load is controllable.
		max_p_mw (float): Maximum active power in megawatts.
		min_p_mw (float): Minimum active power in megawatts.
		max_q_mvar (float): Maximum reactive power in megavars.
		min_q_mvar (float): Minimum reactive power in megavars.
		const_z_percent (float): Constant impedance percentage.
		const_i_percent (float): Constant current percentage.
		sn_mva (float): Apparent power in megavolt-amperes.
		scaling (float): Scaling factor for the load.
		type (str): Type of the load.
	
	Args:
		kwargs: Additional keyword arguments to initialize the load.
	"""
	def __init__(self, kwargs):
		self.p_mw = None
		self.q_mvar = None
		self.controllable = None
		self.max_p_mw = None
		self.min_p_mw = None
		self.max_q_mvar = None
		self.min_q_mvar = None
		self.const_z_percent = None
		self.const_i_percent = None
		self.sn_mva = None
		self.scaling = None
		self.type = None
		super().__init__(**kwargs)
class Transformer(PowerElement):
	"""
	node_p (str or None): Primary node identifier.
	node_s (str or None): Secondary node identifier.
	std_type (str or None): Standard type of the transformer.
	vn_hv_kv (float or None): Nominal voltage on the high voltage side in kV.
	vn_lv_kv (float or None): Nominal voltage on the low voltage side in kV.
	sn_mva (float or None): Nominal apparent power in MVA.
	vk_percent (float or None): Short-circuit voltage in percent.
	vkr_percent (float or None): Short-circuit resistance in percent.
	pfe_kw (float or None): Iron losses in kW.
	i0_percent (float or None): Open-circuit current in percent.
	shift_degree (float or None): Phase shift angle in degrees.
	tap_side (str or None): Tap changer side.
	tap_neutral (int or None): Neutral tap position.
	tap_min (int or None): Minimum tap position.
	tap_max (int or None): Maximum tap position.
	tap_step_percent (float or None): Tap step size in percent.
	tap_step_degree (float or None): Tap step size in degrees.
	tap_phase_shifter (bool or None): Indicates if the transformer is a phase shifter.
	max_loading_percent (float or None): Maximum loading in percent.
	tap_pos (int or None): Current tap position.
	parallel (int or None): Number of parallel transformers.
	df (float or None): Derating factor.
	tap2_pos (int or None): Secondary tap position.
	xn_ohm (float or None): Reactance in ohms.
	tap_dependent_impedance (bool or None): Indicates if the impedance is tap dependent.
	vk_percent_characteristic (list or None): List of short-circuit voltage characteristics.
	vkr_percent_characteristic (list or None): List of short-circuit resistance characteristics.
	vector_group (str or None): Vector group of the transformer.
	vk0_percent (float or None): Zero-sequence short-circuit voltage in percent.
	vkr0_percent (float or None): Zero-sequence short-circuit resistance in percent.
	mag0_percent (float or None): Zero-sequence magnetizing current in percent.
	mag0_rx (float or None): Zero-sequence magnetizing reactance.
	si0_hv_partial (float or None): Zero-sequence impedance on the high voltage side.
	"""
	def __init__(self, kwargs):
		self.node_p = None
		self.node_s = None
		self.std_type = None
		self.vn_hv_kv = None
		self.vn_lv_kv = None
		self.sn_mva = None
		self.vk_percent = None
		self.vkr_percent = None
		self.pfe_kw = None
		self.i0_percent = None
		self.shift_degree = None
		self.tap_side = None
		self.tap_neutral = None
		self.tap_min = None
		self.tap_max = None
		self.tap_step_percent = None
		self.tap_step_degree = None
		self.tap_phase_shifter = None
		self.max_loading_percent = None
		self.tap_pos = None
		self.parallel = None
		self.df = None
		self.tap2_pos = None
		self.xn_ohm = None
		self.tap_dependent_impedance = None
		self.vk_percent_characteristic = None
		self.vkr_percent_characteristic = None
		self.vector_group = None
		self.vk0_percent = None
		self.vkr0_percent = None
		self.mag0_percent = None
		self.mag0_rx = None
		self.si0_hv_partial = None
		super().__init__(**kwargs)

	def update_failure_probability(self, network, intensity=None, ref_return_period=None):
		if self.fragilityCurve == None:
			self.failureProb = None
		else:
			node = network.nodes[self.node_p]
			if intensity == None:
				_, event_intensity = network.event.get_intensity(node.longitude, node.latitude)
				intensity = event_intensity.max()
			if self.return_period != None and ref_return_period != None:
				self.failureProb = network.fragilityCurves[self.fragilityCurve].projected_fc(
					rp = network.returnPeriods[self.return_period],
					ref_rp = network.returnPeriods[ref_return_period],
					xnew = intensity)[0]
			else:
				self.failureProb = network.fragilityCurves[self.fragilityCurve].interpolate(intensity)[0]
class Line(PowerElement):
	'''
	Represents a power line in the network.
	Attributes:
		from_bus (str): The starting bus of the line.
		to_bus (str): The ending bus of the line.
		length_km (float): The length of the line in kilometers.
		r_ohm_per_km (float): The resistance per kilometer in ohms.
		x_ohm_per_km (float): The reactance per kilometer in ohms.
		c_nf_per_km (float): The capacitance per kilometer in nanofarads.
		r0_ohm_per_km (float): The zero-sequence resistance per kilometer in ohms.
		x0_ohm_per_km (float): The zero-sequence reactance per kilometer in ohms.
		c0_nf_per_km (float): The zero-sequence capacitance per kilometer in nanofarads.
		max_i_ka (float): The maximum current in kiloamperes.
		g_us_per_km (float): The conductance per kilometer in microsiemens.
		g0_us_per_km (float): The zero-sequence conductance per kilometer in microsiemens.
		std_type (str): The standard type of the line.
		max_loading_percent (float): The maximum loading percentage.
		df (float): The derating factor.
		parallel (int): The number of parallel lines.
		alpha (float): The temperature coefficient of resistance.
		temperature_degree_celsius (float): The temperature in degrees Celsius.
		tdpf (float): The thermal dynamic performance factor.
		wind_speed_m_per_s (float): The wind speed in meters per second.
		wind_angle_degree (float): The wind angle in degrees.
		conductor_outer_diameter_m (float): The outer diameter of the conductor in meters.
		air_temperature_degree_celsius (float): The air temperature in degrees Celsius.
		reference_temperature_degree_celsius (float): The reference temperature in degrees Celsius.
		solar_radiation_w_per_sq_m (float): The solar radiation in watts per square meter.
		solar_absorptivity (float): The solar absorptivity.
		emissivity (float): The emissivity.
		r_theta_kelvin_per_mw (float): The thermal resistance in Kelvin per megawatt.
		mc_joule_per_m_k (float): The heat capacity in joules per meter per Kelvin.
		lineSpan (float): The span of the line in kilometers.
		type (str): The type of the line.
		failureProb (float): The probability of failure of the line.
	Methods:
		update_failure_probability(network, intensity=None, ref_return_period=None):
			Updates the failure probability of the line based on the network and event intensity.
	'''
	def __init__(self, kwargs):
		self.from_bus = None
		self.to_bus = None
		self.length_km = None
		self.r_ohm_per_km = None
		self.x_ohm_per_km = None
		self.c_nf_per_km = None
		self.r0_ohm_per_km = None
		self.x0_ohm_per_km = None
		self.c0_nf_per_km = None
		self.max_i_ka = None
		self.g_us_per_km = None
		self.g0_us_per_km = None
		self.std_type = None
		self.max_loading_percent = None
		self.df = None
		self.parallel = None
		self.alpha = None
		self.temperature_degree_celsius = None
		self.tdpf = None
		self.wind_speed_m_per_s = None
		self.wind_angle_degree = None
		self.conductor_outer_diameter_m = None
		self.air_temperature_degree_celsius = None
		self.reference_temperature_degree_celsius = None
		self.solar_radiation_w_per_sq_m = None
		self.solar_absorptivity = None
		self.emissivity = None
		self.r_theta_kelvin_per_mw = None
		self.mc_joule_per_m_k = None
		self.lineSpan = None
		self.type = None
		super().__init__(**kwargs)

	def update_failure_probability(self, network, intensity=None, ref_return_period=None):
		if self.fragilityCurve == None:
			self.failureProb = None
		else:
			node1 = network.nodes[self.from_bus]
			node2 = network.nodes[self.to_bus]
			if self.lineSpan == None:
				warnings.warn(f'No lineSpan defined for element {self.id}. Defaulting to 0.2 km.')
				self.lineSpan = 0.2
			if self.length_km == None:
				ValueError(f'No length_km defined for element {self.id}.')
			nb_segments = math.ceil(self.length_km/self.lineSpan)
			probFailure = []
			for i_segment in range(nb_segments+1):
				lon = node1.longitude + i_segment*(node2.longitude-node1.longitude)/nb_segments
				lat = node1.latitude + i_segment*(node2.latitude-node1.latitude)/nb_segments
				if intensity == None:
					_, event_intensity = network.event.get_intensity(lon, lat)
					intensity = event_intensity.max()
				if self.return_period != None and ref_return_period != None:
					probFailure.append(network.fragilityCurves[self.fragilityCurve].projected_fc(
						rp = network.returnPeriods[self.return_period],
						ref_rp = network.returnPeriods[ref_return_period],
						xnew = intensity))
				else:
					probFailure.append(network.fragilityCurves[self.fragilityCurve].interpolate(intensity))		
			self.failureProb = 1-np.prod(1-np.array(probFailure))
class Crew:
	'''
	Represents a crew with specific attributes.

	Attributes:
		id (int): The unique identifier for the crew.
		geodata (GeoData): The geographical data associated with the crew.

	Methods:
		__init__(kwargs): Initializes the Crew instance with given keyword arguments.
	'''
	def __init__(self, kwargs):
		self.id = None
		self.geodata = None
		for key, value in kwargs.items():
			if hasattr(self, key):
				setattr(self, key, value)
			else:
				warnings.warn(f'Input parameter <{key}> unknown in {kwargs}.')
		if self.geodata is None:
			self.geodata = GeoData(0, 0)

### network import from pandapower functions ###

def style_formatting(ws):
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 3)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    font = Font(bold=True)
    border = Border(bottom=Side(border_style="thick"))
    for cell in ws[1]:
        cell.font = font
        cell.border = border

    font = Font(bold=True)
    border = Border(bottom=Side(border_style="thick"))
    for cell in ws[1]:
        cell.font = font
        cell.border = border

def rename_element(sheet, column, values, net, rename_sheet, rename_column, rename = False):
    # TODO: Implement code for rename option!
    if values.empty:
        pass

    elif values.dtype == bool:
        values = values.astype('object').map({True: 'True', False: 'False'})

    elif column == 'name':
        if sheet == 'ln_type':
            values = getattr(net, 'line')['std_type']
        elif sheet == 'tr_type':
            values = getattr(net, 'trafo')['std_type']
        elif rename or values.isna().any() or values.apply(lambda x: isinstance(x, (int, float))).all():
            if sheet == 'nodes':
                for number in values.index:
                    values[number] = 'bus' + str(number + 1)
            else:
                values.reset_index(drop=True, inplace=True)
                for number in values.index:
                    values[number] = rename_sheet[sheet] + str(number + 1)
    
    elif column == 'std_type':
        if values.isna().any():
            values = getattr(net, rename_sheet[sheet])['std_type']
            values.reset_index(drop=True, inplace=True)
            for number in values.index:
                if sheet == 'lines':
                    values[number] = 'line_type' + str(number + 1)
                elif sheet == 'transformers':
                    values[number] = 'trafo_type' + str(number + 1)
            
    elif column in ['node', 'node_p', 'node_s', 'from_bus', 'to_bus']: 
        if isinstance(net, pp.auxiliary.pandapowerNet):              
            bus_column = getattr(net, rename_sheet[sheet])[rename_column[column]]
            bus_names = net.bus.loc[bus_column.tolist(), 'name']
            values = bus_names.reset_index(drop=True)
            values = values.rename(column)
        else:
            raise TypeError('Provided datatype of network is not compliant')
    
    elif column == 'element':
        values = values.astype('object')
        element_type = getattr(net, rename_sheet[sheet])['et']
        for index in range(len(element_type)):
            if element_type.iloc[index] == 'l':
                values.at[index] = getattr(net, rename_sheet['lines'])['name'].loc[values.iloc[index]]
            elif element_type.iloc[index] == 'b':
                values.at[index] = getattr(net, rename_sheet['nodes'])['name'].loc[values.iloc[index]]
            elif element_type.iloc[index] == 't':
                values.at[index] = getattr(net, rename_sheet['transformers'])['name'].loc[values.iloc[index]]
            elif element_type.iloc[index] == 't3':
                values.at[index] = getattr(net, rename_sheet['transformers_3w'])['name'].loc[values.iloc[index]]
            else:
                raise ValueError("Given element type of switch is unknown.")
    
    else:
        pass    
        # print(f"No need to rename for: [{sheet}] - [{column}]") # For debugging
    
    return values

def from_pp(net, profiles = None, rename = False):
    """
	Creates a reXplan compliant network as excel file from pandapower.
    Resilience parameters are not considered in this function.

    INPUT:
		net (dict) - pandapower formatted network
        profiles (dict) - pandapower formatted profiles
        rename (bool) - False: Naming as provided in pandapower net; True: Elements renamed with respective naming of network (under implementation)

    EXAMPLE:
		>>> from_pp(net)
		>>> from_pp(pn.case14(), rename = False)
    """
    # TODO: FOR profiles = VALUE:-----------------------------------
    # TODO: - First column add timesteps -> discuss automation (tab: simulation?)

    # TODO: FOR rename = FALSE:-------------------------------------
    # TODO: - [lines] geodata missing -> first and last location
    # TODO: - [nodes] geodata handling for bus with multiple entries
    # TODO: - Better solution for necessary empty tabs in network.xlsx? -> use keys of rename_column? 

    # TODO: FOR rename = TRUE:--------------------------------------
    # TODO: - Validate same Functionality

    path = os.path.dirname(os.getcwd())
    df_fields_map = pd.read_csv(os.path.join(path + "\\reXplan",'fields_map.csv'))
    rename_column = df_fields_map.set_index('input file field')['pandapower field']
    df_sheets_map = pd.read_csv(os.path.join(path + "\\reXplan",'sheets_map.csv'))
    rename_sheet = df_sheets_map.set_index('input file field')['pandapower field']
    dfs_dict = pd.read_excel('template.xlsx', sheet_name=None)

    for sheet in dfs_dict.keys():

        if sheet == 'cost':
            df_cost = getattr(net, rename_sheet[sheet])
            df_cost = df_cost.sort_values(by='et')

            name_array = np.array([])
            for index in range(len(df_cost.index)):
                if df_cost.iloc[index].et == 'ext_grid':
                    from_sheet = 'external_gen'
                elif df_cost.iloc[index].et == 'gen':
                    from_sheet = 'generators'
                elif df_cost.iloc[index].et == 'sgen':
                    from_sheet = 'static_generators'
                name = dfs_dict[from_sheet].name[df_cost.iloc[index].element] # Extracted name!
                name_array = np.append(name_array, name)

            dfs_dict[sheet] = df_cost
            dfs_dict[sheet]['element'] = name_array         

        elif sheet == 'profiles':
            if profiles:
                if isinstance(profiles, dict):
                    profiles_df = pd.DataFrame()
                    columns_list = pd.DataFrame()
                    field_row = pd.Series(dtype='object')
                    for profile_keys in profiles.keys():
                        if not profiles[profile_keys].empty:
                            df = profiles[profile_keys]
                            value = getattr(net, profile_keys[0])['name']
                            columns_list = pd.concat([columns_list, value], ignore_index=True)
                            profiles_df = pd.concat([profiles_df, df], ignore_index=True, axis=1)
                            field_row_add = pd.Series(["max_" + profile_keys[1]] * len(value), index=range(len(value)))
                            field_row = pd.concat([field_row, field_row_add])
                    if columns_list.T.shape[1] == profiles_df.shape[1]:
                        profiles_df.columns = columns_list.squeeze().tolist()
                    else:
                        raise ValueError ("Length of column name list do not match profile entries.")
                    
                    profiles_df.loc[-1] = pd.Series(dtype='object')
                    profiles_df.index = profiles_df.index + 1
                    profiles_df = profiles_df.sort_index()
                    field_row = field_row.reset_index(drop=True)
                    profiles_df.loc[0] = field_row.values
                    profiles_df = pd.concat([pd.Series('', index=profiles_df.index, name='asset'), profiles_df], axis=1)
                    profiles_df.iloc[0, 0] = "field"

                    dfs_dict[sheet] = profiles_df
                else:
                    raise TypeError('Provided datatype of profiles is not compliant')

        elif sheet == 'ln_type' or sheet == 'tr_type':
            if sheet == 'ln_type':
                var = 'line'
            else:
                var = 'trafo'
            columns = getattr(net, var).columns
            for column in columns:
                    if column in dfs_dict[sheet].keys():
                        old_values = getattr(net, var)[column]
                        values = rename_element(sheet, column, old_values, net, rename_sheet, rename_column, rename)
                        dfs_dict[sheet][column] = values.values # .values?
            dfs_dict[sheet] = dfs_dict[sheet].drop_duplicates()

        else:
            try:
                columns = getattr(net, rename_sheet[sheet]).columns
                for column in columns:
                    if column in dfs_dict[sheet].keys() and not (sheet == 'lines' and column == 'type'):
                        old_values = getattr(net, rename_sheet[sheet])[column]
                        values = rename_element(sheet, column, old_values, net, rename_sheet, rename_column, rename)
                        dfs_dict[sheet][column] = values.values

                    elif column in rename_column.values and (column in dfs_dict[sheet].keys() or column in ['bus', 'hv_bus', 'lv_bus']):
                        column_update = next((key for key, value in rename_column.items() if value == column), None)
                        if column_update is not None:
                            old_values = getattr(net, rename_sheet[sheet])[column]
                            values = rename_element(sheet, column_update, old_values, net, rename_sheet, rename_column, rename)
                            dfs_dict[sheet][column_update] = values.values

                    elif column == 'std_type':
                        old_values = getattr(net, rename_sheet[sheet])[column]
                        values = rename_element(sheet, column, old_values, net, rename_sheet, rename_column, rename)
                        dfs_dict[sheet]['type'] = values.values

                    else:
                        pass
                        # print(f"\nSheet: {rename_sheet[sheet]}; Column: {column} is NOT used in template sheet") # For Debugging
            except:
                # print(f"[{sheet},{column}]") # For Debugging
                pass

    if net.bus_geodata.index.max() == net.bus.index.max():
        net.bus_geodata = net.bus_geodata.sort_index()
        dfs_dict['nodes']['longitude'] = net.bus_geodata.x
        dfs_dict['nodes']['latitude'] = net.bus_geodata.y

    dfs_dict['network']['sn_mva'] = pd.Series(net.sn_mva)
    dfs_dict['network']['f_hz'] = pd.Series(net.f_hz)
    dfs_dict['network']['name'] = pd.Series(net.name)

    wb = Workbook()
    wb.remove(wb['Sheet'])

    nec_sheet_names = ['switches', 'cost', 'tr_type', 'static_generators', 'transformers', 'generators', 'lines']
    for sheet_name, df in dfs_dict.items():
        if not df.empty or sheet_name in nec_sheet_names: 
            ws = wb.create_sheet(sheet_name)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
        else:
            pass
        style_formatting(ws)

    wb.save('network.xlsx')
    network_path = path + r"\jupyter_notebooks\network.xlsx"
    print(f'Network file successfully created: {network_path}')